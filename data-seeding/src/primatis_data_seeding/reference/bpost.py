from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
import xlrd

from primatis_data_seeding.normalization.bpost import normalize_postal_locality


@dataclass(frozen=True)
class BpostLocality:
    postal_code: str
    locality: str


_POSTAL_HEADERS = {"code", "code postal", "postcode", "postal code"}
_LOCALITY_HEADERS = {
    "localite", "locality", "gemeente", "commune", "ville", "plaats"
}


def _fold_header(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text.casefold()).strip()


def _resolve_columns(headers: tuple[object, ...]) -> tuple[int, int]:
    folded = [_fold_header(header) for header in headers]
    postal = [i for i, value in enumerate(folded) if value in _POSTAL_HEADERS]
    locality = [i for i, value in enumerate(folded) if value in _LOCALITY_HEADERS]
    if len(postal) != 1 or len(locality) != 1:
        raise ValueError(
            "Unable to identify exactly one Bpost postal-code column and "
            "one locality column from the workbook header."
        )
    return postal[0], locality[0]


def _coerce_postal_code(value: object) -> object:
    # Legacy .xls files frequently store postal codes as numeric cells.
    # xlrd therefore exposes 1040 as 1040.0. Convert only exact integers.
    if isinstance(value, float):
        if not value.is_integer():
            return value
        integer = int(value)
        if 0 <= integer <= 9999:
            return f"{integer:04d}"
        return str(integer)

    if isinstance(value, int):
        if 0 <= value <= 9999:
            return f"{value:04d}"
        return str(value)

    return value


def _deduplicate(rows) -> list[BpostLocality]:
    result: dict[tuple[str, str], BpostLocality] = {}

    for postal_raw, locality_raw in rows:
        normalized = normalize_postal_locality(
            _coerce_postal_code(postal_raw),
            locality_raw,
        )
        if normalized is None:
            continue

        item = BpostLocality(normalized.postal_code, normalized.locality)
        result[(item.postal_code, item.locality.casefold())] = item

    if not result:
        raise ValueError("Bpost workbook contains no valid postal locality rows.")

    return sorted(
        result.values(),
        key=lambda item: (item.postal_code, item.locality.casefold()),
    )


def _load_xlsx(path: Path) -> list[BpostLocality]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        try:
            headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Bpost workbook is empty.") from exc

        postal_i, locality_i = _resolve_columns(headers)

        return _deduplicate(
            (
                (
                    row[postal_i] if postal_i < len(row) else None,
                    row[locality_i] if locality_i < len(row) else None,
                )
                for row in rows
            )
        )
    finally:
        workbook.close()


def _load_xls(path: Path) -> list[BpostLocality]:
    workbook = xlrd.open_workbook(path)
    if workbook.nsheets == 0:
        raise ValueError("Bpost workbook contains no worksheet.")

    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Bpost workbook is empty.")

    headers = tuple(sheet.cell_value(0, col) for col in range(sheet.ncols))
    postal_i, locality_i = _resolve_columns(headers)

    return _deduplicate(
        (
            (
                sheet.cell_value(row, postal_i),
                sheet.cell_value(row, locality_i),
            )
            for row in range(1, sheet.nrows)
        )
    )


def load_bpost_localities(path: Path) -> list[BpostLocality]:
    if not path.is_file():
        raise ValueError(f"Bpost workbook not found: {path}")

    suffix = path.suffix.casefold()
    if suffix == ".xls":
        return _load_xls(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)

    raise ValueError(
        f"Unsupported Bpost workbook format: {path.suffix!r}. "
        "Expected .xls, .xlsx or .xlsm."
    )
